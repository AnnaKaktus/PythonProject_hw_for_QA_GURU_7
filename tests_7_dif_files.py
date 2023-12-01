import zipfile
import os
from pypdf import PdfReader
from io import TextIOWrapper
from openpyxl import load_workbook
import csv
import pytest

CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIR = os.path.dirname(CURRENT_FILE)
path = os.path.join(CURRENT_DIR, "for_files")
resources = os.path.join(CURRENT_DIR, "resources")
zip_path = os.path.join(resources, "test.zip")


@pytest.fixture
def arc_files():
    if not os.path.exists(resources):
        os.mkdir(resources)
    file_dir = os.listdir(path)
    with zipfile.ZipFile(zip_path, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in file_dir:
            add_file = os.path.join(path, file)
            zf.write(add_file, file)
    yield
    os.remove(zip_path)


def test_create_arch(arc_files):
    assert zipfile.ZipFile("resources/test.zip") != None


def test_open_files(arc_files):
    with zipfile.ZipFile("resources/test.zip") as zip_file:
        # print(zip_file.namelist())
        with zip_file.open("bububu.pdf") as pdf:
            reader = PdfReader(pdf)
            text = reader.pages[0].extract_text()
            assert "бубубу" in text

        with zip_file.open("bububu.csv", "r") as csv_file:
            reader = list(csv.reader(TextIOWrapper(csv_file)))
            # for row in reader:
            # print(', '.join(row))
            assert "bububu" == reader[1][0]

        with zip_file.open("bububu.xlsx") as xlsx:
            workbook = load_workbook(xlsx)
            sheet = workbook.active
            # print(sheet.cell(row=1, column=1))
            assert sheet.cell(row=1, column=1).value == "bububu"
